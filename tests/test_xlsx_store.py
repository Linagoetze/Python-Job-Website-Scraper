"""The xlsx export: a review sheet of unreviewed jobs, an archive sheet of
everything ever seen, and the one place `=HYPERLINK()` formulas are allowed
to exist."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from job_scraper.storage.db import JobStore
from job_scraper.storage.xlsx_store import write_xlsx

# Review sheet column positions (1-based), after the addressable '#' column.
_ROW_NUMBER, _SOURCE, _TITLE, _LOCATION, _DETAIL, _APPLY = 1, 2, 3, 4, 5, 6


def _seed(
    db: Path,
    jobs: list[dict[str, Any]],
    *,
    now: str,
    statuses: dict[str, str] | None = None,
) -> None:
    with JobStore(db) as store:
        run_id = store.begin_run(now)
        store.upsert_jobs(jobs, run_id, now=now)
        for key, status in (statuses or {}).items():
            store.set_status([key], status)
        store.finish_run(run_id)


def _job(key: str, title: str) -> dict[str, str]:
    return {
        "dedupe_key": key,
        "source_name": "acme",
        "title": title,
        "location": "Berlin",
        "detail_url": key,
        "apply_url": "",
    }


def _sheet_rows(xlsx: Path, sheet: str = "Jobs") -> list[tuple]:
    ws = load_workbook(xlsx)[sheet]
    return list(ws.iter_rows(min_row=2, values_only=True))


def _titles(xlsx: Path, sheet: str, column: int) -> list[str]:
    return [row[column - 1] for row in _sheet_rows(xlsx, sheet)]


def test_only_unreviewed_jobs_are_shown(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(
        db,
        [_job("https://x/jobs/1", "Shown"), _job("https://x/jobs/2", "Hidden")],
        now="2026-08-01T00:00:00+00:00",
        statuses={"https://x/jobs/2": "seen"},
    )

    assert write_xlsx(db, xlsx) == 1

    assert _titles(xlsx, "Jobs", _TITLE) == ["Shown"]


def test_show_all_puts_every_job_on_the_review_sheet(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(
        db,
        [
            _job("https://x/jobs/1", "New one"),
            _job("https://x/jobs/2", "Reviewed"),
            _job("https://x/jobs/3", "Gone"),
        ],
        now="2026-08-01T00:00:00+00:00",
        statuses={"https://x/jobs/2": "rejected", "https://x/jobs/3": "delisted"},
    )

    assert write_xlsx(db, xlsx, show_all=True) == 3

    rows = _sheet_rows(xlsx, "Jobs")
    assert sorted(r[_TITLE - 1] for r in rows) == ["Gone", "New one", "Reviewed"]
    # The status column exists only in this mode, where it carries information.
    assert sorted(r[-1] for r in rows) == ["delisted", "new", "rejected"]
    assert load_workbook(xlsx)["Jobs"].cell(row=1, column=7).value == "status"


def test_the_archive_sheet_holds_every_job_whatever_its_status(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(
        db,
        [
            _job("https://x/jobs/1", "Unreviewed"),
            _job("https://x/jobs/2", "Rejected"),
            _job("https://x/jobs/3", "Shortlisted"),
            _job("https://x/jobs/4", "Delisted"),
        ],
        now="2026-08-01T00:00:00+00:00",
        statuses={
            "https://x/jobs/2": "rejected",
            "https://x/jobs/3": "shortlisted",
            "https://x/jobs/4": "delisted",
        },
    )

    assert write_xlsx(db, xlsx) == 1, "the review sheet still shows only the unreviewed one"

    archive = _sheet_rows(xlsx, "Archive")
    assert sorted(r[2] for r in archive) == [
        "Delisted",
        "Rejected",
        "Shortlisted",
        "Unreviewed",
    ], "nothing is hidden from the owner"
    assert sorted(r[0] for r in archive) == ["delisted", "new", "rejected", "shortlisted"]
    assert all(r[4] and r[5] for r in archive), "first_seen and last_seen are shown"


def test_the_row_number_column_matches_the_worksheet_row(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(
        db,
        [_job(f"https://x/jobs/{i}", f"Job {i}") for i in range(3)],
        now="2026-08-01T00:00:00+00:00",
    )

    write_xlsx(db, xlsx)

    ws = load_workbook(xlsx)["Jobs"]
    assert ws.cell(row=1, column=_ROW_NUMBER).value == "#"
    numbers = [ws.cell(row=r, column=_ROW_NUMBER).value for r in range(2, ws.max_row + 1)]
    assert numbers == [2, 3, 4], "the printed number is the row Excel shows in its gutter"


def test_the_export_records_which_job_each_row_holds(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(
        db,
        [_job("https://x/jobs/1", "First"), _job("https://x/jobs/2", "Second")],
        now="2026-08-01T00:00:00+00:00",
    )

    write_xlsx(db, xlsx)

    ws = load_workbook(xlsx)["Jobs"]
    with JobStore(db) as store:
        mapping = store.export_row_map()
    assert len(mapping) == 2
    for row_number, key in mapping.items():
        assert ws.cell(row=row_number, column=_ROW_NUMBER).value == row_number
        assert ws.cell(row=row_number, column=_DETAIL).value == f'=HYPERLINK("{key}")'


def test_a_new_export_replaces_the_previous_row_mapping(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(db, [_job(f"https://x/jobs/{i}", f"Job {i}") for i in range(3)],
          now="2026-08-01T00:00:00+00:00")
    write_xlsx(db, xlsx)

    with JobStore(db) as store:
        store.set_status(["https://x/jobs/0", "https://x/jobs/1"], "seen")
    write_xlsx(db, xlsx)

    with JobStore(db) as store:
        assert store.export_row_map() == {2: "https://x/jobs/2"}, "only the last export addressable"


def test_urls_become_hyperlink_formulas_at_export_time(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(db, [_job("https://x/jobs/1", "Analyst")], now="2026-08-01T00:00:00+00:00")

    write_xlsx(db, xlsx)

    (row,) = _sheet_rows(xlsx)
    assert row[_DETAIL - 1] == '=HYPERLINK("https://x/jobs/1")', "plain URL in, formula out"
    assert row[_APPLY - 1] is None, "no apply URL, no formula"


def test_rows_from_the_two_most_recent_runs_are_highlighted(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(db, [_job("https://x/jobs/old", "Old")], now="2026-08-01T00:00:00+00:00")
    _seed(db, [_job("https://x/jobs/mid", "Mid")], now="2026-08-02T00:00:00+00:00")
    _seed(db, [_job("https://x/jobs/new", "New")], now="2026-08-03T00:00:00+00:00")

    write_xlsx(db, xlsx)

    ws = load_workbook(xlsx)["Jobs"]
    filled = {
        ws.cell(row=r, column=_TITLE).value
        for r in range(2, ws.max_row + 1)
        if str(ws.cell(row=r, column=_SOURCE).fill.start_color.rgb or "").endswith("C6EFCE")
    }
    assert filled == {"Mid", "New"}


def test_empty_store_writes_an_empty_table(tmp_path: Path) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    assert write_xlsx(db, xlsx) == 0
    assert _sheet_rows(xlsx) == []
    assert _sheet_rows(xlsx, "Archive") == []


def test_a_failed_save_leaves_the_previous_spreadsheet_intact(
    tmp_path: Path, monkeypatch: Any
) -> None:
    db, xlsx = tmp_path / "jobs.sqlite3", tmp_path / "jobs.xlsx"
    _seed(db, [_job("https://x/jobs/1", "First")], now="2026-08-01T00:00:00+00:00")
    write_xlsx(db, xlsx)
    before = xlsx.read_bytes()

    _seed(db, [_job("https://x/jobs/2", "Second")], now="2026-08-02T00:00:00+00:00")
    from openpyxl import Workbook

    def boom(self: Workbook, path: str) -> None:
        raise OSError("disk full")

    monkeypatch.setattr(Workbook, "save", boom)
    try:
        write_xlsx(db, xlsx)
    except OSError:
        pass

    assert xlsx.read_bytes() == before, "the readable spreadsheet survived"
    assert list(tmp_path.glob(".jobs-*.xlsx")) == [], "no temp file left behind"
    with JobStore(db) as store:
        assert len(store.export_row_map()) == 1, "row numbers still address the file on disk"
