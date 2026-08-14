"""Write a styled xlsx: a review sheet of unreviewed jobs and an archive of
every job ever seen, highlighting recent rows in green.

The one place Excel syntax is allowed to exist: the store holds plain URLs,
and the `=HYPERLINK()` formulas are generated here at export time.

The review sheet's first column is the row number the review commands address
(`python -m job_scraper.review --reject 7`). It is written into the sheet
rather than left implicit in Excel's gutter because sorting the table in Excel
rearranges the rows under the gutter numbers but carries the `#` column along
with its job.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from job_scraper.storage.db import JobStore

_ROW_FIELD = "#"
_SCORE_FIELD = "score"
_DISPLAY_FIELDS = [
    "source_name",
    "title",
    "location",
    "score",
    "score_reasoning",
    "score_flags",
    "detail_url",
    "apply_url",
]
# The status column earns its place only when the sheet can hold more than one
# status: in the default view every row says 'new'.
_STATUS_FIELD = "status"
_ARCHIVE_FIELDS = [
    "status",
    "source_name",
    "title",
    "location",
    "first_seen",
    "last_seen",
    "detail_url",
    "apply_url",
]
_URL_FIELDS = {"detail_url", "apply_url"}

_REVIEW_SHEET = "Jobs"
_ARCHIVE_SHEET = "Archive"

_HEADER_FONT = Font(bold=True)
_NEW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_LINK_FONT = Font(color="0563C1", underline="single")
_COL_WIDTHS: dict[str, float] = {
    _ROW_FIELD: 5,
    "status": 12,
    "source_name": 18,
    "title": 42,
    "location": 26,
    "score": 7,
    "score_reasoning": 60,
    "score_flags": 30,
    "first_seen": 22,
    "last_seen": 22,
    "detail_url": 55,
    "apply_url": 45,
}


def _hyperlink_formula(url: str) -> str:
    """Single-arg HYPERLINK so the formula has no commas."""
    return '=HYPERLINK("' + url.replace('"', '""') + '")'


def _write_header(ws: Worksheet, fields: list[str]) -> None:
    ws.freeze_panes = "A2"
    for col, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = _HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = _COL_WIDTHS.get(field, 20)


def _write_cell(ws: Worksheet, ws_row: int, col: int, field: str, value: Any) -> None:
    cell = ws.cell(row=ws_row, column=col)
    if field == _SCORE_FIELD:
        # Numeric so Excel sorts it as a number; NULL (unscored) stays blank,
        # which keeps 0 — a real score — distinguishable from "not scored".
        if value is not None:
            cell.value = int(value)
        return
    raw = str(value or "")
    if field in _URL_FIELDS and raw:
        cell.value = _hyperlink_formula(raw)
        cell.font = _LINK_FONT
    else:
        cell.value = raw


def _sorted_for_review(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Best score first; unscored rows last, grouped by source, newest first.

    The stable sorts make source/recency the tiebreak among equal scores, so
    the pre-WP7 ordering survives within each score band — and entirely, when
    scoring was skipped and every score is NULL.
    """
    rows.sort(key=lambda r: str(r.get("first_seen") or ""), reverse=True)
    rows.sort(key=lambda r: str(r.get("source_name") or "").lower())
    rows.sort(key=lambda r: r["score"] if r.get("score") is not None else -1, reverse=True)
    return rows


def _sorted_for_archive(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Newest first — the archive is history, and history reads backwards."""
    rows.sort(key=lambda r: str(r.get("source_name") or "").lower())
    rows.sort(key=lambda r: str(r.get("first_seen") or ""), reverse=True)
    return rows


def _save_atomically(wb: Workbook, xlsx_path: Path) -> None:
    """Save via a temp file in the same directory, then os.replace().

    A crash or a locked file mid-save must not leave a truncated spreadsheet
    where a readable one used to be.
    """
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=xlsx_path.parent, prefix=".jobs-", suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise


def write_xlsx(db_path: Path, xlsx_path: Path, *, show_all: bool = False) -> int:
    """Write the review sheet and the archive sheet. Returns the review row count.

    The review sheet holds unreviewed ('new') jobs only — what the owner opens
    the file to see — unless *show_all* is set, in which case it holds every
    job and gains a status column. The archive sheet always holds every job
    ever seen, whatever its status, so nothing is ever hidden.

    Rows first seen in the two most recent storing runs are filled light
    green. Every job stored by one run shares a single first_seen timestamp,
    so the two most recent distinct first_seen values among the displayed rows
    identify those runs.

    Writing the file and recording which job each review row holds happen in
    one store transaction: if the save fails, the previous export stays
    addressable rather than the row numbers advancing past a file the owner
    never received.
    """
    with JobStore(db_path) as store:
        archive_rows = _sorted_for_archive(store.all_jobs())
        review_rows = _sorted_for_review(
            list(archive_rows) if show_all else store.jobs_with_status(("new",))
        )

        recent_seen = sorted(
            {str(r.get("first_seen") or "") for r in review_rows}, reverse=True
        )[:2]

        review_fields = [_ROW_FIELD, *_DISPLAY_FIELDS] + ([_STATUS_FIELD] if show_all else [])

        wb = Workbook()
        ws = wb.active
        ws.title = _REVIEW_SHEET
        _write_header(ws, review_fields)

        export_rows: list[tuple[int, str]] = []
        for row_idx, row in enumerate(review_rows):
            ws_row = row_idx + 2  # 1-based, offset by the header
            export_rows.append((ws_row, str(row["dedupe_key"])))
            is_new = str(row.get("first_seen") or "") in recent_seen

            for col, field in enumerate(review_fields, start=1):
                if field == _ROW_FIELD:
                    # A number, so it stays readable next to Excel's own gutter.
                    ws.cell(row=ws_row, column=col, value=ws_row)
                else:
                    _write_cell(ws, ws_row, col, field, row.get(field))
                if is_new:
                    ws.cell(row=ws_row, column=col).fill = _NEW_FILL

        archive = wb.create_sheet(_ARCHIVE_SHEET)
        _write_header(archive, _ARCHIVE_FIELDS)
        for row_idx, row in enumerate(archive_rows):
            for col, field in enumerate(_ARCHIVE_FIELDS, start=1):
                _write_cell(archive, row_idx + 2, col, field, row.get(field))

        _save_atomically(wb, xlsx_path)
        store.replace_export_rows(export_rows)

    return len(review_rows)
